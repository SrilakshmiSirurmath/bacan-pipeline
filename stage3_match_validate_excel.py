# stage3_match_validate_excel.py
#
# Pipeline:
# 1) Read extracted texts (out/invoice_text.txt, out/ead_text.txt)
# 2) Build SAFE AI prompts (redact + trim)
# 3) AI extract invoice + EAD rows
# 4) ALSO deterministically extract EAD "Denominazione di origine" per progressivo from raw text
# 5) Normalize rows into dicts
# 6) Match invoice lines -> EAD lines
# 7) Validate shipment-level + line-level (customs-grade checks)
# 8) Write packing_list.xlsx (fallback to CSV if openpyxl missing) + issues.json

from __future__ import annotations

from pathlib import Path
import json
import re
from collections import Counter
from rapidfuzz import fuzz 
from openpyxl import load_workbook
from io import BytesIO

import pandas as pd

from stage2b_ai_extract_openai import ai_extract_invoice, ai_extract_ead
from stage1b_redact_trim import redact, trim_invoice_text, trim_ead_text

def extract_invoice_compliance(invoice_text: str) -> dict:
    t = invoice_text or ""

    out = {
        "supplier_vat": None,           # P.IVA
        "supplier_cod_fisc": None,      # Cod.Fisc. (often same line)
        "supplier_eori": None,          # Codice EORI (exporter)
        "supplier_rex": None,           # Numero REX
        "consignee_eori": None,         # Codice Eori destinatario
        "incoterm": None,               # Incoterms = EXW ...
        "total_colli": None,            # N.ro Colli
        "gross_kg": None,               # Peso lordo KG
        "net_kg": None,                 # peso netto kg
        "pallet_count": None,           # "posti su 2 pallet"
    }

    # VAT / Cod.Fisc / P.IVA (simple heuristic)
    m = re.search(r"\bCod\.?Fisc\.?\s*e\s*P\.?Iva\s+([0-9]+)", t, re.IGNORECASE)
    if m:
        out["supplier_vat"] = m.group(1).strip()

    # Consignee EORI (destinatario)
    m = re.search(r"\bCodice\s+Eori\s+destinatario\s*=\s*([A-Z0-9]+)", t, re.IGNORECASE)
    if m:
        out["consignee_eori"] = m.group(1).strip()

    # Incoterm
    m = re.search(r"\bIncoterms?\s*=\s*([A-Z]{3})", t, re.IGNORECASE)
    if m:
        out["incoterm"] = m.group(1).upper().strip()

    # Supplier REX + EORI (exporter)
    m = re.search(r"\bNumero\s+Rex\s+([A-Z0-9]+)", t, re.IGNORECASE)
    if m:
        out["supplier_rex"] = m.group(1).strip()

    m = re.search(r"\bCodice\s+EORI\s+([A-Z0-9]+)", t, re.IGNORECASE)
    if m:
        out["supplier_eori"] = m.group(1).strip()

    # Colli / weights (you already have something similar, keep yours if preferred)
    m = re.search(r"\bN\.?ro\s+Colli\b\s*([0-9\.\,]+)", t, re.IGNORECASE)
    if m:
        out["total_colli"] = parse_int_loose(m.group(1))

    m = re.search(r"\bPeso\s+Lordo\b.*?\bKg\b\.?\s*([0-9\.\,]+)", t, re.IGNORECASE)
    if m:
        out["gross_kg"] = parse_float_locale(m.group(1))

    m = re.search(r"\bpeso\s+netto\b.*?\bkg\b\.?\s*([0-9\.\,]+)", t, re.IGNORECASE)
    if m:
        out["net_kg"] = parse_float_locale(m.group(1))

    # pallets (basic)
    m = re.search(r"\bposti\s+su\s+(\d+)\s+pallet", t, re.IGNORECASE)
    if m:
        out["pallet_count"] = int(m.group(1))

    return out

def country_from_denom(denom: str) -> str:
    d = (denom or "").upper()

    italy_tokens = ["DOC", "DOCG", "IGT", "IGP", "SICILIA", "SICILIANE", "ITALIA", "NERO D'AVOLA", "TERRE"]
    if any(t in d for t in italy_tokens):
        return "Italy"

    if any(t in d for t in ["AOC", "AOP", "BORDEAUX", "BOURGOGNE", "CHAMPAGNE"]):
        return "France"

    if any(t in d for t in ["DOCA", "DOCa", "DO", "RIOJA", "RIBERA"]):
        return "Spain"

    return ""

def shipper_name_from_ead_text(ead_text: str) -> str | None:
    t = ead_text or ""
    # keep line structure
    lines = [ln.strip() for ln in t.splitlines() if ln.strip()]

    for i, ln in enumerate(lines):
        if re.search(r"\(2\.b\)\s*Nome\s+dello\s+speditore\s*:", ln, re.IGNORECASE):
            # what comes after colon (often just "AGRICOLA")
            parts = ln.split(":", 1)
            tail = parts[1].strip() if len(parts) > 1 else ""

            # previous line often contains the start of company name
            prev = lines[i-1] if i > 0 else ""

            # Combine safely
            full = f"{prev} {tail}".strip()
            full = re.sub(r"\s+", " ", full)

            return full if full else None

    return None


def build_customs_excel(matches, template_path: str, inv_ai, ead_text) -> bytes:
    df = build_output_df(matches)

    wb = load_workbook(template_path)
    ws = wb.active

    shipper_name = ship_name = shipper_name_from_ead_text(ead_text)

    # --- Fill INVOICE LEVEL DATA ---
    ws["C3"] = shipper_name
    ws["C4"] = inv_ai.supplier_eori
    ws["C5"] = inv_ai.supplier_rex
    ws["C6"] = inv_ai.incoterm

    start_row = 14  # where items start in template

    # --- Fill item rows ---
    for idx, row in df.iterrows():
        r = start_row + idx

        pieces = None
        if row["CASES / COLLI"] and row["BOTTLES PER CASE"]:
            pieces = int(row["CASES / COLLI"]) * int(row["BOTTLES PER CASE"])

        ws[f"A{r}"] = idx + 1
        ws[f"B{r}"] = row["DESCRIPTION"]
        ws[f"C{r}"] = row["CUSTOMS COMMODITY CODE"]
        ws[f"D{r}"] = row["% ALCOHOL"]
        ws[f"E{r}"] = pieces
        ws[f"F{r}"] = row["GROSS WEIGHT (KG)"]
        ws[f"G{r}"] = row["NET WEIGHT (KG)"]
        ws[f"H{r}"] = row["INVOICE VALUE (EUR)"]
        ws[f"I{r}"] = country_from_denom(row["DENOMINAZIONE DI ORIGINE"])

    # --- Clear unused template rows ---
    last_filled_row = start_row + len(df) - 1
    max_template_rows = start_row + 20  # buffer

    for r in range(last_filled_row + 1, max_template_rows):
        for col in ["A","B","C","D","E","F","G","H","I"]:
            ws[f"{col}{r}"] = None

    # --- Dynamic TOTAL row ---
    total_row = start_row + len(df) + 1

    ws[f"C{total_row}"] = "TOTAL"
    ws[f"E{total_row}"] = f"=SUM(E{start_row}:E{total_row-2})"
    ws[f"F{total_row}"] = f"=SUM(F{start_row}:F{total_row-2})"
    ws[f"G{total_row}"] = f"=SUM(G{start_row}:G{total_row-2})"
    ws[f"H{total_row}"] = f"=SUM(H{start_row}:H{total_row-2})"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ------------------------------------------------------------
# Similarity: rapidfuzz (if installed) or lightweight fallback
# ------------------------------------------------------------

def token_set_ratio(a: str, b: str) -> float:
        a = (a or "").strip()
        b = (b or "").strip()
        if not a or not b:
            return 0.0
        return float(fuzz.token_set_ratio(a, b))


# -----------------------------
# Helpers
# -----------------------------
def is_num(x):
    return isinstance(x, (int, float)) and x is not None


def safe_float(x):
    try:
        return float(x)
    except Exception:
        return None


def parse_int_loose(x):
    if x is None:
        return None
    s = str(x).strip()
    s = re.sub(r"[^\d]", "", s)
    return int(s) if s else None


def parse_float_locale(x):
    """
    Parses numbers like:
      1.702,00  -> 1702.00  (IT)
      1702,00   -> 1702.00
      1702.00   -> 1702.00
    """
    if x is None:
        return None
    s = str(x).strip().replace(" ", "")
    if "." in s and "," in s:
        # thousand '.' and decimal ','
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def normalize_bottle_liters(val):
    """
    Normalize bottle size to liters.
      - 750 -> 0.75 (ml)
      - 75 -> 0.75 (cl) if it looks like cl
    """
    if val is None:
        return None
    try:
        v = float(val)
    except Exception:
        return None

    if v > 10:
        return v / 1000.0

    if 1 < v <= 100 and abs(v - round(v)) < 1e-9:
        if v in (75, 70, 50):
            return v / 100.0

    return v


def bottles_per_case_from_desc(desc: str):
    """
    Extract bottles-per-case from invoice description in many languages.
    Examples:
      - "CRT DA 6 BTLS"
      - "CASE OF 6 BOTTLES"
      - "CARTON 6 BOUTEILLES"
      - "IN CRT DA 6"
    """
    if not desc:
        return None

    patterns = [
        r"\b(?:CRT|CARTON|CARTONE|CASE)\s*(?:DA|DI|OF)?\s*(\d+)\s*(?:BTLS|BOTTLES?|BOTTIGLIE|BOUTEILLES|BT)\b",
        r"\bCASE\s+OF\s+(\d+)\s+BOTTLES?\b",
    ]
    for p in patterns:
        m = re.search(p, desc, re.IGNORECASE)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return None
    return None


def liters_from_invoice(inv_line):
    """
    Preferred liters calc:
      cases * bottles_per_case * bottle_liters

    Fallback:
      bottles_total * bottle_liters (ONLY if bottles_total seems like a count)
    """
    cases = inv_line.get("cases")
    bpc = inv_line.get("bottles_per_case")
    bl = normalize_bottle_liters(inv_line.get("bottle_liters"))

    if cases is not None and bpc is not None and bl is not None:
        try:
            return float(cases) * float(bpc) * float(bl)
        except Exception:
            pass

    bt = inv_line.get("bottles_total")
    if bt is not None and bl is not None:
        try:
            # guard: prevent “750,000 next to BT” trap
            if float(bt) <= 5000:
                return float(bt) * float(bl)
        except Exception:
            pass

    return None


# ------------------------------------------------------------
# Deterministic extraction from raw texts (for audits & origin)
# ------------------------------------------------------------
def extract_invoice_totals(invoice_text: str) -> dict:
    t = invoice_text or ""
    out = {
        "invoice_total_colli": None,
        "invoice_gross_kg": None,
        "invoice_net_kg": None,
    }

    m = re.search(r"\bN\.?ro\s+Colli\b\s*([0-9\.\,]+)", t, re.IGNORECASE)
    if m:
        out["invoice_total_colli"] = parse_int_loose(m.group(1))

    m = re.search(r"\bPeso\s+lordo\b.*?\bKG\b[:\s]*([0-9\.\,]+)", t, re.IGNORECASE)
    if m:
        out["invoice_gross_kg"] = parse_float_locale(m.group(1))

    m = re.search(r"\bpeso\s+netto\b.*?\bkg\b\.?[:\s]*([0-9\.\,]+)", t, re.IGNORECASE)
    if m:
        out["invoice_net_kg"] = parse_float_locale(m.group(1))

    return out


def extract_ead_packaging_colli_sum(ead_text: str) -> int | None:
    t = ead_text or ""
    colli = re.findall(r"\bNumero\s+di\s+colli\b:\s*([0-9\.\,]+)", t, re.IGNORECASE)
    if not colli:
        return None
    vals = [parse_int_loose(x) for x in colli]
    vals = [v for v in vals if v is not None]
    return sum(vals) if vals else None

# -----------------------------
# Normalizers (AI -> dict lines)
# -----------------------------
def normalize_invoice_rows(ai_invoice) -> dict:
    lines = []
    for r in ai_invoice.rows:
        desc = (r.description or "").strip()
        bpc = r.bottles_per_case
        if bpc is None:
            bpc = bottles_per_case_from_desc(desc)

        lines.append({
            "description": desc,
            "cn_code": r.cn_code,
            "abv_percent": safe_float(r.abv_percent),
            "bottles": r.bottles,
            "bottle_liters": normalize_bottle_liters(r.bottle_liters),
            "bottles_total": r.bottles_total,
            "cases": r.cases,
            "bottles_per_case": bpc,
            "invoice_value_eur": safe_float(r.invoice_value_eur),
            "lot": r.lot,
        })

    return {
        "invoice_number": getattr(ai_invoice, "invoice_number", None),
        "invoice_date": getattr(ai_invoice, "invoice_date", None),
        "arc": getattr(ai_invoice, "arc", None),
        "incoterm": getattr(ai_invoice, "incoterm", None),
        "lines": lines,
    }


def normalize_ead_rows(ai_ead) -> dict:

    lines = []
    for idx, r in enumerate(ai_ead.rows, start=1):
        prog = r.progressivo if r.progressivo is not None else idx
        prog = int(prog)

        desc = (getattr(r, "description", "") or "").strip()
        extra = (getattr(r, "designazione_commerciale", "") or "").strip()

        if extra and extra.lower() not in desc.lower():
            designazione_commerciale = f"{desc} {extra}".strip()
        else:
            designazione_commerciale = desc        
        
        designation = getattr(r, "designation", None) or getattr(r, "description", None) or ""
        denominazione_origine = getattr(r, "denominazione_origine", None)


        lines.append({
            "progressivo": prog,
            "cn_code": r.cn_code,
            "abv_percent": safe_float(r.abv_percent),
            "ead_liters": safe_float(r.ead_liters),
            "ead_gross_kg": safe_float(r.ead_gross_kg),
            "ead_net_kg": safe_float(r.ead_net_kg),
            "designation": re.sub(r"\s+", " ", (designation or "")).strip(),
            "denominazione_origine": re.sub(r"\s+", " ", denominazione_origine).strip() if denominazione_origine else None,
            "designazione_commerciale": re.sub(r"\s+", " ", (designazione_commerciale or "")).strip() if designazione_commerciale else None,
            "cases": r.cases,  # often None per product line
        })

    return {
        "arc": getattr(ai_ead, "arc", None),
        "invoice_number": getattr(ai_ead, "invoice_number", None),
        "invoice_date": getattr(ai_ead, "invoice_date", None),
        "lines": lines,
    }


# -----------------------------
# Matching
# -----------------------------
def match_invoice_to_ead(inv_lines, ead_lines):
    used = set()
    matches = []

    for inv in inv_lines:
        best = None
        best_score = -1

        for ead in ead_lines:
            if ead["progressivo"] in used:
                continue

            # Hard filter: CN code if present on both
            if inv.get("cn_code") and ead.get("cn_code"):
                if str(inv["cn_code"]).strip() != str(ead["cn_code"]).strip():
                    continue

            score = 0

            # Liters strongest
            inv_liters = liters_from_invoice(inv)
            if inv_liters is not None and ead.get("ead_liters") is not None:
                diff = abs(inv_liters - ead["ead_liters"])
                if diff <= 0.5:
                    score += 80
                elif diff <= 2.0:
                    score += 50
                else:
                    score -= 50

            # Cases (if EAD provides per-line cases)
            if inv.get("cases") is not None and ead.get("cases") is not None:
                try:
                    score += 40 if int(inv["cases"]) == int(ead["cases"]) else -30
                except Exception:
                    pass

            # ABV soft
            if inv.get("abv_percent") is not None and ead.get("abv_percent") is not None:
                try:
                    diff_abv = abs(float(inv["abv_percent"]) - float(ead["abv_percent"]))
                    if diff_abv <= 0.3:
                        score += 25
                    elif diff_abv <= 0.7:
                        score += 10
                    else:
                        score -= 25
                except Exception:
                    pass

            # Description tie-breaker
            inv_name = (inv.get("description") or "").strip()
            ead_name = (ead.get("designation") or "").strip()
            if inv_name and ead_name:
                score += token_set_ratio(inv_name, ead_name) / 2.0

            if score > best_score:
                best_score = score
                best = ead

        if best:
            used.add(best["progressivo"])
        matches.append((inv, best, best_score))

    return matches


# -----------------------------
# Validation (Line-level + Shipment-level)
# -----------------------------
def validate_lines(matches, *, liters_tol=0, abv_tol_warn=0, abv_tol_fail=0, name_warn_threshold=100.0):
    """
    Classes:
      - PRODUCT_IDENTITY_CHECK
      - QUANTITY_INTEGRITY_CHECK
      - COMPLETENESS_CHECK
    """
    issues = []

    def add(check_class, issue_type, severity, **kwargs):
        issues.append({"check_class": check_class, "type": issue_type, "severity": severity, **kwargs})

    for inv, ead, score in matches:
        inv_desc = (inv.get("description") or "").strip()

        # Must have match
        if ead is None:
            add("PRODUCT_IDENTITY_CHECK", "NO_MATCH", "FAIL", invoice_desc=inv_desc, match_score=score)
            continue

        # Completeness: invoice per-row mandatory
        required_inv = ["description", "cn_code", "abv_percent", "cases", "bottles_per_case", "bottle_liters"]
        for f in required_inv:
            v = inv.get(f)
            if v is None or (isinstance(v, str) and not v.strip()):
                sev = "FAIL" if f in ("cases", "bottles_per_case", "bottle_liters", "cn_code") else "WARN"
                add("COMPLETENESS_CHECK", "MISSING_INVOICE_FIELD", sev, invoice_desc=inv_desc, missing_field=f)

        # Completeness: EAD liters
        if ead.get("ead_liters") is None:
            add("COMPLETENESS_CHECK", "MISSING_EAD_LITERS", "FAIL", invoice_desc=inv_desc, ead_progressivo=ead.get("progressivo"))

        # CN mismatch
        inv_cn = inv.get("cn_code")
        ead_cn = ead.get("cn_code")
        if inv_cn and ead_cn and str(inv_cn).strip() != str(ead_cn).strip():
            add("PRODUCT_IDENTITY_CHECK", "CN_CODE_MISMATCH", "FAIL",
                invoice_desc=inv_desc, invoice_cn_code=inv_cn, ead_cn_code=ead_cn)

        # ABV mismatch
        inv_abv = safe_float(inv.get("abv_percent"))
        ead_abv = safe_float(ead.get("abv_percent"))
        if is_num(inv_abv) and is_num(ead_abv):
            diff_abv = abs(inv_abv - ead_abv)
            if diff_abv > abv_tol_fail:
                add("PRODUCT_IDENTITY_CHECK", "ABV_MISMATCH", "FAIL",
                    invoice_desc=inv_desc, invoice_abv=inv_abv, ead_abv=ead_abv, diff=diff_abv)
            elif diff_abv > abv_tol_warn:
                add("PRODUCT_IDENTITY_CHECK", "ABV_MISMATCH", "WARN",
                    invoice_desc=inv_desc, invoice_abv=inv_abv, ead_abv=ead_abv, diff=diff_abv)

        # Bottle size sanity
        bl = safe_float(inv.get("bottle_liters"))
        if bl is not None and (bl < 0.05 or bl > 5.0):
            add("PRODUCT_IDENTITY_CHECK", "BOTTLE_SIZE_SUSPICIOUS", "WARN", invoice_desc=inv_desc, bottle_liters=bl)

        # Liters invariant
        inv_liters = liters_from_invoice(inv)
        ead_liters = ead.get("ead_liters")
        if is_num(inv_liters) and is_num(ead_liters):
            if abs(inv_liters - ead_liters) > liters_tol:
                add("QUANTITY_INTEGRITY_CHECK", "LITERS_MISMATCH", "FAIL",
                    invoice_desc=inv_desc, invoice_calc_liters=inv_liters, ead_liters=ead_liters)

        # Cases mismatch (only if EAD has per-line cases)
        inv_cases = inv.get("cases")
        ead_cases = ead.get("cases")
        if inv_cases is not None and ead_cases is not None:
            try:
                if int(inv_cases) != int(ead_cases):
                    add("QUANTITY_INTEGRITY_CHECK", "CASES_MISMATCH", "FAIL",
                        invoice_desc=inv_desc, invoice_cases=inv_cases, ead_cases=ead_cases)
            except Exception:
                add("QUANTITY_INTEGRITY_CHECK", "CASES_PARSE_ERROR", "WARN",
                    invoice_desc=inv_desc, invoice_cases=inv_cases, ead_cases=ead_cases)

        # Weight sanity
        gross = safe_float(ead.get("ead_gross_kg"))
        net = safe_float(ead.get("ead_net_kg"))
        if is_num(gross) and is_num(net) and gross <= net:
            add("QUANTITY_INTEGRITY_CHECK", "WEIGHT_GROSS_LE_NET", "WARN",
                invoice_desc=inv_desc, ead_gross_kg=gross, ead_net_kg=net)

        if is_num(net) and is_num(ead_liters):
            if abs(net - ead_liters) > max(2.0, 0.02 * ead_liters):
                add("QUANTITY_INTEGRITY_CHECK", "NETKG_LITERS_SUSPICIOUS", "WARN",
                    invoice_desc=inv_desc, ead_net_kg=net, ead_liters=ead_liters)

        # Name similarity (warn only)
        inv_name = (inv.get("description") or "").strip()
        ead_name = (ead.get("description") or "").strip()
        if inv_name and ead_name:
            sim = token_set_ratio(inv_name, ead_name)
            if sim < name_warn_threshold:
                add("PRODUCT_IDENTITY_CHECK", "LOW_NAME_SIMILARITY", "WARN",
                    invoice_desc=inv_desc, similarity=sim, invoice_name=inv_name, ead_name=ead_name)

        # Lot presence (warn only)
        if inv.get("lot") is None:
            add("PRODUCT_IDENTITY_CHECK", "MISSING_LOT", "WARN", invoice_desc=inv_desc)

        # Denominazione di origine (you explicitly want it)
        if ead.get("denominazione_origine") is None:
            add("COMPLETENESS_CHECK", "MISSING_DENOMINAZIONE_ORIGINE", "WARN",
                invoice_desc=inv_desc, ead_progressivo=ead.get("progressivo"))

    return issues


def validate_shipment(inv_ai, ead_ai, inv_lines, ead_lines, *, invoice_text: str, ead_text: str):
    """
    Shipment-level checks:
      - DOCUMENT_CONSISTENCY_CHECK: ARC / invoice number / invoice date
      - QUANTITY_INTEGRITY_CHECK: total liters, total cases (if present), total colli (invoice vs EAD packaging sum)
      - COMPLETENESS_CHECK: mandatory headers
      - Packaging sanity: total bottles / total colli approx bottles_per_case mode
    """
    issues = []

    def add(check_class, issue_type, severity, **kwargs):
        issues.append({"check_class": check_class, "type": issue_type, "severity": severity, **kwargs})
     inv_comp = extract_invoice_compliance(invoice_text)
   #ead_comp = extract_ead_compliance(ead_text)

    # ---- Compliance completeness checks (Invoice) ----
    mandatory_invoice = [
        ("supplier_vat", "MISSING_SUPPLIER_VAT"),
        ("supplier_eori", "MISSING_SUPPLIER_EORI"),
        ("supplier_rex", "MISSING_SUPPLIER_REX"),
        ("consignee_eori", "MISSING_CONSIGNEE_EORI"),
        ("incoterm", "MISSING_INCOTERM"),
        ("total_colli", "MISSING_TOTAL_COLLI"),
        ("gross_kg", "MISSING_GROSS_KG"),
        ("net_kg", "MISSING_NET_KG"),
    ]
    for key, issue_type in mandatory_invoice:
        if not inv_comp.get(key):
            add("COMPLETENESS_CHECK", issue_type, "FAIL")

    # Pallets: often required but sometimes absent -> warn (your choice)
    if not inv_comp.get("pallet_count"):
        add("COMPLETENESS_CHECK", "MISSING_PALLET_COUNT", "WARN")
        
    inv_no = getattr(inv_ai, "invoice_number", None)
    ead_no = getattr(ead_ai, "invoice_number", None)
    inv_date = getattr(inv_ai, "invoice_date", None)
    ead_date = getattr(ead_ai, "invoice_date", None)
    inv_arc = getattr(inv_ai, "arc", None)
    ead_arc = getattr(ead_ai, "arc", None)

    # Completeness (headers)
    if not inv_no:
        add("COMPLETENESS_CHECK", "MISSING_INVOICE_NUMBER", "FAIL")
    if not ead_no:
        add("COMPLETENESS_CHECK", "MISSING_EAD_INVOICE_NUMBER", "FAIL")
    if not inv_arc:
        add("COMPLETENESS_CHECK", "MISSING_ARC_IN_INVOICE", "FAIL")
    if not ead_arc:
        add("COMPLETENESS_CHECK", "MISSING_ARC_IN_EAD", "FAIL")

    # Document consistency
    if inv_arc and ead_arc and str(inv_arc).strip() != str(ead_arc).strip():
        add("DOCUMENT_CONSISTENCY_CHECK", "ARC_MISMATCH", "FAIL", invoice_arc=inv_arc, ead_arc=ead_arc)

    if inv_no and ead_no and str(inv_no).strip() != str(ead_no).strip():
        add("DOCUMENT_CONSISTENCY_CHECK", "INVOICE_NUMBER_MISMATCH", "FAIL", invoice_number=inv_no, ead_invoice_number=ead_no)

    if inv_date and ead_date and str(inv_date).strip() != str(ead_date).strip():
        add("DOCUMENT_CONSISTENCY_CHECK", "INVOICE_DATE_MISMATCH", "WARN", invoice_date=str(inv_date), ead_invoice_date=str(ead_date))

    # Totals: liters
    inv_liters_vals = []
    for l in inv_lines:
        lit = liters_from_invoice(l)
        if lit is not None:
            inv_liters_vals.append(float(lit))
    inv_liters_sum = sum(inv_liters_vals) if inv_liters_vals else None

    ead_liters_vals = []
    for l in ead_lines:
        lit = liters_from_invoice(l)
        if lit is not None:
            ead_liters_vals.append(float(lit))
    ead_liters_sum = sum(ead_liters_vals) if ead_liters_vals else None

    if inv_liters_sum is not None and ead_liters_sum is not None:
        if abs(inv_liters_sum - ead_liters_sum) > max(1.0, 0.005 * ead_liters_sum):
            add("QUANTITY_INTEGRITY_CHECK", "TOTAL_LITERS_MISMATCH", "FAIL",
                invoice_calc_liters_total=inv_liters_sum, ead_liters_total=ead_liters_sum)

    # Totals: cases (only if EAD has per-line cases)
    inv_cases_sum = sum(int(l.get("cases") or 0) for l in inv_lines)
    ead_cases_sum = sum(int(l.get("cases") or 0) for l in ead_lines)

    if inv_cases_sum and ead_cases_sum is not None and inv_cases_sum != ead_cases_sum:
        add("QUANTITY_INTEGRITY_CHECK", "TOTAL_CASES_MISMATCH", "FAIL",
            invoice_sum_cases=inv_cases_sum, ead_sum_cases=ead_cases_sum)

    # Totals: EAD weight sanity
    ead_gross_vals = [float(l.get("ead_gross_kg")) for l in ead_lines if l.get("ead_gross_kg") is not None]
    ead_net_vals = [float(l.get("ead_net_kg")) for l in ead_lines if l.get("ead_net_kg") is not None]
    ead_gross_sum = sum(ead_gross_vals) if ead_gross_vals else None
    ead_net_sum = sum(ead_net_vals) if ead_net_vals else None
    if ead_gross_sum is not None and ead_net_sum is not None and ead_gross_sum <= ead_net_sum:
        add("QUANTITY_INTEGRITY_CHECK", "TOTAL_GROSS_LE_NET", "WARN",
            ead_gross_total=ead_gross_sum, ead_net_total=ead_net_sum)

    # NEW: invoice total colli vs sum(EAD Numero di colli) from packaging section
    inv_meta = extract_invoice_totals(invoice_text)
    inv_total_colli = inv_meta.get("invoice_total_colli")
    ead_colli_sum = extract_ead_packaging_colli_sum(ead_text)

    if inv_total_colli is not None and ead_colli_sum is not None:
        if int(inv_total_colli) != int(ead_colli_sum):
            add("QUANTITY_INTEGRITY_CHECK", "TOTAL_COLLI_MISMATCH", "FAIL",
                invoice_total_colli=inv_total_colli, ead_packaging_colli_sum=ead_colli_sum)

    # NEW: packaging sanity: total bottles / total colli ≈ bottles_per_case mode
    total_bottles = 0
    bpcs = []
    for l in inv_lines:
        c = l.get("cases")
        bpc = l.get("bottles_per_case")
        if bpc is not None:
            try:
                bpcs.append(int(bpc))
            except Exception:
                pass
        if c is not None and bpc is not None:
            try:
                total_bottles += int(c) * int(bpc)
            except Exception:
                pass

    if inv_total_colli is not None and inv_total_colli > 0 and total_bottles > 0:
        bottles_per_carton = total_bottles / float(inv_total_colli)

        if bpcs:
            mode_bpc = Counter(bpcs).most_common(1)[0][0]
            if abs(bottles_per_carton - mode_bpc) > 0.25:
                add("QUANTITY_INTEGRITY_CHECK", "PACKAGING_UNIT_MISMATCH", "FAIL",
                    total_bottles=total_bottles,
                    invoice_total_colli=inv_total_colli,
                    bottles_per_carton=bottles_per_carton,
                    mode_bottles_per_case=mode_bpc)
        else:
            add("QUANTITY_INTEGRITY_CHECK", "PACKAGING_UNIT_UNCHECKED", "WARN",
                reason="No bottles_per_case extracted from invoice rows")

    return issues


def build_output_df(matches):
    rows = []
    for inv, ead, score in matches:
        rows.append({
            # Prefer EAD description if present (designazione commerciale is the cleanest)
            "DESCRIPTION": (
                (ead.get("designazione_commerciale") if ead else None)
            ),

            "DENOMINAZIONE DI ORIGINE": (ead.get("denominazione_origine") if ead else None),

            # EAD first
            "CUSTOMS COMMODITY CODE": (ead.get("cn_code") if ead else None) or inv.get("cn_code"),

            # EAD first
            "% ALCOHOL": (
                (ead.get("abv_percent") if ead else None)
                if (ead and ead.get("abv_percent") is not None)
                else inv.get("abv_percent")
            ),

            # EAD first
            "CASES / COLLI": (ead.get("cases") if ead else None) or inv.get("cases"),

            # Packaging detail usually only invoice has
            "BOTTLES PER CASE": inv.get("bottles_per_case"),
            "BOTTLE SIZE (L)": normalize_bottle_liters(inv.get("bottle_liters")),

            "TOTAL LITERS (calc)": liters_from_invoice(inv),
            "EAD LITERS": ead.get("ead_liters") if ead else None,

            "GROSS WEIGHT (KG)": ead.get("ead_gross_kg") if ead else None,
            "NET WEIGHT (KG)": ead.get("ead_net_kg") if ead else None,

            # This MUST come from invoice (EAD doesn't carry item value)
            "INVOICE VALUE (EUR)": inv.get("invoice_value_eur"),

            "LOT": inv.get("lot"),
            "EAD PROGRESSIVO": ead.get("progressivo") if ead else None,
            "MATCH_SCORE": score,
        })
    return pd.DataFrame(rows)


def write_excel_or_csv(df: pd.DataFrame, out_xlsx: Path) -> tuple[Path, str]:
    """
    Write Excel if openpyxl is available; else fallback to CSV.
    Returns (written_path, message)
    """
    try:
        import openpyxl  # noqa: F401
        df.to_excel(out_xlsx, index=False)
        return out_xlsx, "xlsx"
    except Exception as e:
        out_csv = out_xlsx.with_suffix(".csv")
        df.to_csv(out_csv, index=False)
        return out_csv, f"csv (Excel write failed: {type(e).__name__}: {e})"


# -----------------------------
# Main
# -----------------------------
def main():
    out_dir = Path("out")
    out_dir.mkdir(exist_ok=True)

    invoice_text = (out_dir / "invoice_text.txt").read_text()
    ead_text = (out_dir / "ead_text.txt").read_text()

    # SAFE prompts for AI
    inv_safe = redact(invoice_text)
    ead_safe = redact(ead_text)

    # AI extraction
    inv_ai = ai_extract_invoice(inv_safe)
    ead_ai = ai_extract_ead(ead_safe)

    # Normalize
    inv = normalize_invoice_rows(inv_ai)
    ead = normalize_ead_rows(ead_ai)

    # Shipment-level checks
    shipment_issues = validate_shipment(
        inv_ai, ead_ai, inv["lines"], ead["lines"],
        invoice_text=invoice_text, ead_text=ead_text
    )

    # Matching + line-level checks
    matches = match_invoice_to_ead(inv["lines"], ead["lines"])
    line_issues = validate_lines(matches)

    issues = shipment_issues + line_issues

    # Output
    df = build_output_df(matches)
    out_xlsx = out_dir / "packing_list.xlsx"
    written_path, mode = write_excel_or_csv(df, out_xlsx)

    (out_dir / "issues.json").write_text(json.dumps(issues, indent=2, ensure_ascii=False))

    print("✅ Wrote:", written_path, f"[{mode}]")
    print("✅ Wrote:", out_dir / "issues.json")
    print("Issues:", len(issues))
    if issues:
        for i in issues[:12]:
            print("-", i["severity"], i["type"], i.get("invoice_desc", ""))


if __name__ == "__main__":
    main()
